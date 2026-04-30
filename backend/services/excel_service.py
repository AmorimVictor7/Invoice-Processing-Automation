# Serviço de geração do arquivo Excel de tabulação das NFs.
# Usa openpyxl para criar a planilha com formatação visual (cores, bordas, larguras de coluna).

from __future__ import annotations
import os
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.invoice import InvoiceData


# Cabeçalhos das colunas na ordem em que aparecem na planilha
HEADERS = [
    "Data Processamento",
    "Fornecedor",
    "Nº Invoice",
    "Data Emissão",
    "Descrição",
    "Moeda",
    "Subtotal",
    "Impostos",
    "Valor Original",
    "Cotação (R$)",
    "Valor Convertido (R$)",
    "Centro de Custo",
    "Observações",
    "Arquivo Vinculado",
]

# Estilos visuais da planilha
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # azul escuro no cabeçalho
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)  # texto branco bold
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")       # azul claro nas linhas pares (zebra)
BORDER_SIDE = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE
)

# Larguras das colunas em caracteres (mesma ordem que HEADERS)
COL_WIDTHS = [20, 25, 18, 14, 40, 8, 14, 14, 16, 14, 22, 18, 30, 45]


def _renamed_file(inv: InvoiceData) -> str:
    """
    Gera o nome padronizado do arquivo para a coluna 'Arquivo Vinculado'.
    Formato: FORNECEDOR_NUMINVOICE_DATA_VALOR+MOEDA.ext
    Ex: GOOGLE_INV-001_20240315_250USD.pdf
    """
    supplier = (inv.supplier or "DESCONHECIDO").upper().replace(" ", "_")
    number = (inv.invoice_number or "SEM_NUM").upper().replace(" ", "_").replace("/", "_")
    date = (inv.issue_date or "SEM_DATA").replace("-", "")
    amount = ""
    if inv.total_amount is not None:
        amount = f"_{inv.total_amount:.0f}{inv.currency or ''}"
    ext = Path(inv.original_filename).suffix or ".pdf"
    return f"{supplier}_{number}_{date}{amount}{ext}"


def generate_excel(invoices: List[InvoiceData], output_path: str) -> str:
    """
    Cria o arquivo Excel com todas as NFs confirmadas.
    - Linha 1: cabeçalho fixo com fundo azul escuro e texto branco
    - Linhas 2+: dados das NFs com zebra (linhas pares em azul claro)
    - Colunas numéricas (7–11): formatadas com separador de milhar
    - Linha 1 congelada (freeze_panes) para facilitar scroll
    - Filtro automático habilitado em todas as colunas
    Retorna o caminho do arquivo salvo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Escreve e formata a linha de cabeçalho
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30  # altura maior para o cabeçalho

    processed_at = datetime.now().strftime("%Y-%m-%d %H:%M")

    for row_idx, inv in enumerate(invoices, start=2):
        # Linhas pares recebem fundo azul claro; ímpares ficam sem preenchimento
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()

        # Calcula o valor convertido em BRL: prioriza total_amount * exchange_rate,
        # depois usa converted_amount se já veio preenchido do frontend
        converted = None
        if inv.total_amount is not None and inv.exchange_rate:
            converted = inv.total_amount * inv.exchange_rate
        elif inv.converted_amount is not None:
            converted = inv.converted_amount

        # Dados na mesma ordem que HEADERS
        row_data = [
            processed_at,
            inv.supplier or "",
            inv.invoice_number or "",
            inv.issue_date or "",
            inv.description or "",
            inv.currency or "",
            inv.subtotal,
            inv.taxes,
            inv.total_amount,
            inv.exchange_rate,
            converted,
            inv.cost_center or "",
            inv.observations or "",
            _renamed_file(inv),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            # Aplica formato numérico nas colunas de valores monetários (colunas 7 a 11)
            if col_idx in (7, 8, 9, 10, 11) and value is not None:
                cell.number_format = "#,##0.00"

    # Ajusta a largura de cada coluna conforme COL_WIDTHS
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Congela a linha 1 para que o cabeçalho fique visível ao rolar verticalmente
    ws.freeze_panes = "A2"

    # Habilita filtro automático em toda a tabela (usuário pode filtrar por fornecedor, moeda, etc.)
    ws.auto_filter.ref = ws.dimensions

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path